import argparse
import json
import csv
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter

def load_config(config_path):
    with open(config_path, 'r') as f:
        return json.load(f)

def create_barcode(product_code, barcode_type='code39'):
    barcode_class = barcode.get_barcode_class(barcode_type)
    barcode_instance = barcode_class(product_code, writer=ImageWriter(), add_checksum=False)
    barcode_path = Path('temp_barcode')
    barcode_instance.save(barcode_path)
    return barcode_path.with_suffix('.png')

def create_product_label(product_name, product_code, save_path, config):
    image = Image.new('RGB', tuple(config['image_size']), color=config['background_color'])
    draw = ImageDraw.Draw(image)
    
    title_font = ImageFont.truetype(config['title_font'], config['title_font_size'])
    content_font = ImageFont.truetype(config['content_font'], config['content_font_size'])
    
    # Draw border
    if config['draw_border']:
        draw.rectangle([config['border_padding'], config['border_padding'], 
                        config['image_size'][0] - config['border_padding'], 
                        config['image_size'][1] - config['border_padding']], 
                       outline=config['border_color'], width=config['border_width'])
    
    # Draw content
    for field in config['fields']:
        draw.text(tuple(field['position']), field['label'], font=title_font, fill=field['color'])
        content = product_name if field['type'] == 'product_name' else product_code
        draw.text((field['position'][0], field['position'][1] + field['content_offset']), 
                  content, font=content_font, fill=field['color'])
    
    # Create and insert barcode
    barcode_path = create_barcode(product_code, config['barcode_type'])
    with Image.open(barcode_path) as barcode_img:
        barcode_img = barcode_img.resize(tuple(config['barcode_size']))
        image.paste(barcode_img, tuple(config['barcode_position']))
    
    barcode_path.unlink()  # Remove temporary barcode file
    
    image.save(save_path)

def process_file(file_path, save_folder, config):
    save_folder = Path(save_folder)
    save_folder.mkdir(exist_ok=True)
    
    file_extension = Path(file_path).suffix.lower()
    
    try:
        if file_extension == '.xlsx':
            process_excel(file_path, save_folder, config)
        elif file_extension == '.csv':
            process_csv(file_path, save_folder, config)
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
    except Exception as e:
        print(f"An error occurred while processing the file: {e}")

def process_excel(excel_path, save_folder, config):
    with load_workbook(excel_path) as wb:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            create_label_from_row(row, save_folder, config)

def process_csv(csv_path, save_folder, config):
    with open(csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        csv_reader = csv.reader(csvfile)
        next(csv_reader)  # Skip header row
        for row in csv_reader:
            create_label_from_row(row, save_folder, config)

def create_label_from_row(row, save_folder, config):
    product_name, product_code = row[0], row[1]
    if product_name and product_code:
        product_code = product_code.strip()  # Loại bỏ khoảng trắng đầu và cuối
        save_path = save_folder / f"{product_code}.png"
        create_product_label(product_name, product_code, save_path, config)
        print(f"Created label for product: {product_name}, Code: {product_code}")
    else:
        print(f"Skipped invalid row: {row}")

def main(file_path, save_folder, config):
    process_file(file_path, save_folder, config)

if __name__ == "__main__":
    main()
